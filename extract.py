import openpyxl, json, re, unicodedata, datetime, os, shutil
from collections import defaultdict, Counter

BASE_XLSX = "/Users/joaquinbanchio/Desktop/Trabajo Papi/Analisis suelos"
SRC_FILE = "UNIFICADO_v8 (no coprimido y sin pocos datos).xlsx"
OUT_DIR = os.path.dirname(os.path.abspath(__file__))

def campana_fallback_date(campana):
    """Rows collapsed under Excel row-grouping (see read_sheet) often have no
    Fecha of their own, only a Campaña like '2025/2026' or '25/26'. Falling
    back to Jul 1 of the campaign's first year keeps them plottable on the
    evolution chart, at campaign-level granularity instead of an exact date."""
    if not campana:
        return None
    m = re.match(r"^(\d{2}|\d{4})/", str(campana).strip())
    if not m:
        return None
    yr = m.group(1)
    yr = ("20" + yr) if len(yr) == 2 else yr
    return f"{yr}-07-01"

def strip_accents(s):
    if s is None: return ""
    s = str(s)
    nfkd = unicodedata.normalize('NFKD', s)
    return "".join(c for c in nfkd if not unicodedata.combining(c))

NULL_TEXT = {"", "—", "-", "–", "sin fecha", "s/f", "s/d"}

def clean(v):
    if v is None: return None
    if isinstance(v, str):
        v = v.strip()
        if v.lower() in NULL_TEXT: return None
        return v
    if isinstance(v, datetime.datetime):
        return v.strftime("%Y-%m-%d")
    return v

def clean_date(v):
    """Always normalize to canonical ISO YYYY-MM-DD, regardless of whether the
    source cell was a real Excel date or text typed as DD/MM/YYYY (both occur
    in this workbook -- some rows are DD/MM/YYYY text, which sorts wrong
    lexicographically if left as-is: '16/06/2025' < '2022-05-27')."""
    v = clean(v)
    if v is None: return None
    if isinstance(v, str):
        m = re.match(r"^(\d{2})/(\d{2})/(\d{4})$", v)
        if m:
            dd, mm, yyyy = m.groups()
            return f"{yyyy}-{mm}-{dd}"
        m = re.match(r"^(\d{4})-(\d{2})-(\d{2})$", v)
        if m:
            return v
        return None
    return v

def to_num(v):
    v = clean(v)
    if v is None: return None
    if isinstance(v, (int, float)): return v
    try:
        return float(str(v).replace(",", "."))
    except Exception:
        return None

def norm_prof(v):
    v = clean(v)
    if v is None: return None
    s = str(v).strip().upper().replace(" ", "")
    s = s.replace("CM", "")
    return s + " cm" if s else None

STOP = {"LOTE","LT","CAMPO","ZONA","DE","LA","EL","LOS","LAS","L","HA","CM","UNICO","ÚNICO"}
TYPE_TOKENS = {"BIO","BIOLOGICO","BIOLOGICA","QUIM","QUIMICO","QUIMICA","QCO","Q","TESTIGO","HUMIC"}
TIPO_DISPLAY = {"Biologico": "Biológico", "Quimico": "Químico", "Testigo": "Testigo"}

TYPO_FIX = [
    (r"\bANGER\b", "ANGEL"),
    (r"\b23 EL TACA\b", "EL TALA 23"),
    (r"\bEL TACA\b", "EL TALA"),
]

def normalize_core(name):
    if not name: return "", None
    s = strip_accents(name).upper()
    s = re.sub(r"[\.,/]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    for pat, rep in TYPO_FIX:
        s = re.sub(pat, rep, s)
    s = re.sub(r"\bL(\d)", r"\1", s)
    tokens = s.split(" ")
    tipo_found = None
    kept = []
    for t in tokens:
        tt = t.strip()
        if tt in TYPE_TOKENS:
            if tt in ("BIO", "BIOLOGICO", "BIOLOGICA"): tipo_found = TIPO_DISPLAY["Biologico"]
            elif tt in ("QUIM", "QUIMICO", "QUIMICA", "QCO", "Q"): tipo_found = TIPO_DISPLAY["Quimico"]
            elif tt == "TESTIGO": tipo_found = TIPO_DISPLAY["Testigo"]
            elif tt == "HUMIC": tipo_found = TIPO_DISPLAY["Biologico"]
            continue
        if tt in STOP:
            continue
        if tt == "":
            continue
        kept.append(tt)
    core = " ".join(kept).strip()
    if core == "":
        core = s
    return core, tipo_found

def sig_tokens(core):
    return set(t for t in core.split(" ") if t and (t.isdigit() or len(t) >= 3))

PROD_ALIASES = {
    "DIEGO BANCHIO": "Banchio Diego",
    "BANCHIO DIEGO": "Banchio Diego",
    "DIEGO FORNI": "Diego Forni",
    "LUCAS BAUDINO": "Lucas Baudino",
    "ANGEL BOSCHETTO": "Boschetto Angel",
    "BOSCHETTO ANGEL": "Boschetto Angel",
}

def norm_prod(name):
    if not name: return "Sin dato"
    s = strip_accents(name).upper().strip()
    s = re.sub(r"\s+", " ", s)
    return PROD_ALIASES.get(s, name.strip())

# ---------- READ SOURCE (two separate sheets, kept as two separate datasets) ----------
# The workbook itself keeps "Laboratorios" (Molisol/Clínica/AFA/AgLab -- all
# conventional chemistry) and "TecnoSustrato" (microbiological) as distinct
# sheets. That split is intentional: these are two different lab
# methodologies, and this script never merges a conventional record with a
# microbiological one, at any step, for any field.
wb = openpyxl.load_workbook(f"{BASE_XLSX}/{SRC_FILE}", data_only=True)

CHEM_PARAM_MAP = [
    ("pH", "pH"), ("CE (µS/cm)", "CE"), ("MO (%)", "MO"), ("C Org. (%)", "C_Org"),
    ("N Total (%)", "N_Total"), ("N-NO3 (ppm)", "N_NO3"), ("P Bray I (ppm)", "P_BrayI"),
    ("P Total (ppm)", "P_Total"), ("Azufre S (ppm)", "Azufre_S"), ("CIC (meq/100g)", "CIC"),
    ("Ca Intercamb. (ppm)", "Ca_Interc"), ("Mg Intercamb. (ppm)", "Mg_Interc"),
    ("K Intercamb. (ppm)", "K_Interc"), ("Na Intercamb. (ppm)", "Na_Interc"),
    ("PSI (%)", "PSI"), ("Sat. Bases (%)", "Sat_Bases"), ("Zn (ppm)", "Zn"),
    ("Mn (ppm)", "Mn"), ("Fe (ppm)", "Fe"), ("Cu (ppm)", "Cu"), ("Bo (ppm)", "Bo"),
    ("Agua Útil (%)", "Agua_Util"),
]
BIO_PARAM_MAP = [
    ("pH", "pH"), ("CE (µS/cm)", "CE"),
    ("N bio (ppm)", "N_bio"), ("P bio (ppm)", "P_bio"), ("K bio (ppm)", "K_bio"),
    ("CO2 Resp. (ppm)", "CO2_Resp"), ("% Orgánico", "Pct_Organico"), ("% Mineral", "Pct_Mineral"),
    ("% Centro Gaseoso", "Pct_CentroGaseoso"), ("% SBE", "Pct_SBE"), ("ICS (1–3)", "ICS"),
    ("Act. Biológica", "Act_Biologica"), ("Humif. MO", "Humif_MO"), ("Miner. MO", "Miner_MO"),
    ("SBE (bio)", "SBE"), ("Compactación", "Compactacion"), ("Respiración", "Respiracion"),
]

def read_sheet(sheet_name, dataset, param_map):
    ws = wb[sheet_name]
    headers = [ws.cell(row=3, column=c).value for c in range(1, ws.max_column + 1)]
    idx = {h: i + 1 for i, h in enumerate(headers) if h}
    out_rows = []
    n_hidden_skipped = 0
    n_fallback_date = 0
    for r in range(4, ws.max_row + 1):
        # The sheet uses Excel row-grouping: a lot's individual per-depth (or
        # per-muestra, for TecnoSustrato) rows are collapsed/hidden under a
        # visible "▶ PROMEDIO" summary row. Per instruction, hidden rows are
        # dropped entirely -- only what's visible in the sheet counts.
        if ws.row_dimensions[r].hidden:
            n_hidden_skipped += 1
            continue

        lab = clean(ws.cell(row=r, column=idx["Laboratorio"]).value)
        if not lab or lab == "Laboratorio":
            continue
        lote_raw = clean(ws.cell(row=r, column=idx["Lote"]).value)
        if not lote_raw:
            continue

        prod_raw = clean(ws.cell(row=r, column=idx["Productor/Propietario"]).value)
        core, tipo_from_lote = normalize_core(lote_raw)
        tipo_col = clean(ws.cell(row=r, column=idx["Manejo/Tipo"]).value)
        tipo = tipo_col or tipo_from_lote

        params = {}
        for xlname, key in param_map:
            params[key] = to_num(ws.cell(row=r, column=idx[xlname]).value)

        campana = clean(ws.cell(row=r, column=idx["Campaña"]).value)
        fecha = clean_date(ws.cell(row=r, column=idx["Fecha"]).value)
        fecha_is_estimate = False
        if fecha is None:
            fallback = campana_fallback_date(campana)
            if fallback:
                fecha = fallback
                fecha_is_estimate = True
                n_fallback_date += 1

        out_rows.append({
            "dataset": dataset,
            "lab": lab,
            "informe": clean(ws.cell(row=r, column=idx["Archivo/Informe"]).value),
            "productor": norm_prod(prod_raw),
            "localidad": clean(ws.cell(row=r, column=idx["Municipio/Localidad"]).value),
            "lote": lote_raw,
            "lote_core": core,
            "tipo": tipo,
            "muestra": clean(ws.cell(row=r, column=idx["Posición/Muestra"]).value),
            "prof": norm_prof(ws.cell(row=r, column=idx["Profundidad"]).value),
            "fecha": fecha,
            "fecha_is_estimate": fecha_is_estimate,
            "campana": campana,
            "params": params,
            "score": to_num(ws.cell(row=r, column=idx["Score ponderado"]).value) if "Score ponderado" in idx else None,
            "clasificacion": clean(ws.cell(row=r, column=idx["Clasificación ponderada"]).value) if "Clasificación ponderada" in idx else None,
        })
    print(f"  {sheet_name}: skipped {n_hidden_skipped} hidden rows, {n_fallback_date} rows use a campaign-start estimate for Fecha")
    return out_rows

chem_rows = read_sheet("Laboratorios", "quimico", CHEM_PARAM_MAP)
bio_rows = read_sheet("TecnoSustrato", "biologico", BIO_PARAM_MAP)

# Manual corrections requested directly (not present in the source sheet).
# Each entry pins the exact row so nothing else can be touched by accident.
for r in chem_rows:
    if (r["lote"] == "CAPILLA" and r["productor"] == "Banchio Diego"
            and r["fecha"] == "2026-06-17" and r["lab"] == "AFA Planta Formuladora"
            and r["tipo"] == "Químico"):
        r["tipo"] = None

# ---------- MANUAL ENTRIES ----------
# Analyses sent to us directly as a standalone PDF, not (yet) in the master
# Excel. Kept in their own file and re-applied on every extraction so a
# future "here's the new Excel" doesn't silently drop them.
MANUAL_ENTRIES_PATH = os.path.join(OUT_DIR, "manual_entries.json")

def load_manual_entries():
    if not os.path.isfile(MANUAL_ENTRIES_PATH):
        return []
    with open(MANUAL_ENTRIES_PATH, encoding="utf-8") as f:
        entries = json.load(f)
    out = []
    for e in entries:
        lote_raw = e["lote_raw"]
        core, tipo_from_lote = normalize_core(lote_raw)
        out.append({
            "dataset": e["dataset"],
            "lab": e["lab"],
            "informe": e.get("informe"),
            "informe_url": e.get("informe_url"),
            "productor": norm_prod(e.get("productor_raw")),
            "localidad": e.get("localidad"),
            "lote": lote_raw,
            "lote_core": core,
            "tipo": e.get("manejo_tipo") or tipo_from_lote,
            "muestra": e.get("muestra"),
            "prof": norm_prof(e.get("profundidad_raw")),
            "fecha": clean_date(e.get("fecha_raw")),
            "fecha_is_estimate": False,
            "campana": e.get("campana"),
            "params": dict(e.get("params") or {}),
            "score": e.get("score"),
            "clasificacion": e.get("clasificacion"),
        })
    return out

manual_rows = load_manual_entries()
chem_rows += [r for r in manual_rows if r["dataset"] == "quimico"]
bio_rows += [r for r in manual_rows if r["dataset"] == "biologico"]
print("manual entries loaded:", len(manual_rows))

rows = chem_rows + bio_rows

n_chem = len(chem_rows)
n_bio = len(bio_rows)
print("chem rows:", n_chem, "| bio rows:", n_bio, "| total:", len(rows))

# ---------- LINK SOURCE PDF REPORTS ----------
# "informe" is a mix of real report codes (S23-0102 B), real PDF filenames
# (Banchio 526 (1).pdf), and -- for many rows -- just the lot name again,
# which is NOT a document reference. Only link when a PDF's filename
# uniquely contains the informe string, and never when the only candidate is
# a climate report that happens to also mention the lot name.
REPORTS_SRC_EXCLUDE_DIRS = {"panel-suelos-web"}
REPORTS_EXCLUDE_KEYWORDS = ["climatol", "clima"]
REPORTS_OUT_DIR = os.path.join(OUT_DIR, "reports")

def _norm_fname(s):
    return re.sub(r"[\s_\-]+", " ", s).strip().lower()

all_pdfs = []
for root, dirs, files in os.walk(BASE_XLSX):
    if any(ex in root for ex in REPORTS_SRC_EXCLUDE_DIRS):
        continue
    for f in files:
        if f.lower().endswith(".pdf"):
            all_pdfs.append(os.path.join(root, f))

informe_to_path = {}
for r in rows:
    inf = r.get("informe")
    if not inf or inf in informe_to_path:
        continue
    inf_n = _norm_fname(inf)
    candidates = [p for p in all_pdfs if inf_n in _norm_fname(os.path.basename(p))]
    candidates = [p for p in candidates if not any(k in _norm_fname(os.path.basename(p)) for k in REPORTS_EXCLUDE_KEYWORDS)]
    if len(candidates) == 1:
        informe_to_path[inf] = candidates[0]

if os.path.isdir(REPORTS_OUT_DIR):
    shutil.rmtree(REPORTS_OUT_DIR)
os.makedirs(REPORTS_OUT_DIR, exist_ok=True)
informe_to_slug = {}
for i, (inf, src_path) in enumerate(sorted(informe_to_path.items())):
    slug = f"report_{i:03d}.pdf"
    shutil.copyfile(src_path, os.path.join(REPORTS_OUT_DIR, slug))
    informe_to_slug[inf] = f"reports/{slug}"

for r in rows:
    r["informe_file"] = r.get("informe_url") or informe_to_slug.get(r.get("informe"))

print("linked source reports:", len(informe_to_slug), "of", len(set(r["informe"] for r in rows if r.get("informe"))), "distinct informes")

# ---------- GROUPING (dataset is part of the key -- never merged across datasets) ----------
def group_key(row):
    return (row["dataset"], strip_accents(row["productor"]).upper(), row["lote_core"])

groups = defaultdict(list)
for row in rows:
    groups[group_key(row)].append(row)

print("initial groups (pre-merge):", len(groups))

# ---------- WITHIN-DATASET MERGE (consolidate typos/relabeling of the same
# physical lot across labs of the SAME dataset -- e.g. Molisol vs AFA both
# reporting "COL 3"). Two keys from different datasets are never compared. ----------
def tokens_for_group(key):
    return sig_tokens(key[2])

all_keys = list(groups.keys())
doc_freq = defaultdict(int)
for k in all_keys:
    for t in tokens_for_group(k):
        doc_freq[t] += 1
DISTINCTIVE_MAX_DF = 3
NEVER_DISTINCTIVE = {"ESTE", "OESTE", "NORTE", "SUR", "CENTRO", "MITAD"}

def distinctive(toks):
    # a bare number is never enough evidence on its own -- lot numbers like
    # "5", "23", "50" recur constantly across unrelated fields/producers, so
    # only alphabetic words (rarer, more specific) count as distinctive
    return {t for t in toks if not t.isdigit() and doc_freq[t] <= DISTINCTIVE_MAX_DF and t not in NEVER_DISTINCTIVE}

def has_digit(tok):
    return any(ch.isdigit() for ch in tok)

def same_operator(prod_a, prod_b):
    a, b = prod_a.strip(), prod_b.strip()
    if not a or not b: return False
    short, long_ = (a, b) if len(a) <= len(b) else (b, a)
    return short in long_ and len(short) >= 5

class UnionFind:
    def __init__(self, keys):
        self.parent = {k: k for k in keys}
    def find(self, x):
        while self.parent[x] != x:
            self.parent[x] = self.parent[self.parent[x]]
            x = self.parent[x]
        return x
    def union(self, a, b):
        ra, rb = self.find(a), self.find(b)
        if ra != rb:
            self.parent[ra] = rb

uf = UnionFind(all_keys)
merge_log = []
for i in range(len(all_keys)):
    for j in range(i + 1, len(all_keys)):
        ak, bk = all_keys[i], all_keys[j]
        if ak[0] != bk[0]:
            continue  # different dataset (quimico vs biologico) -- never merge
        if uf.find(ak) == uf.find(bk):
            continue
        atoks, btoks = tokens_for_group(ak), tokens_for_group(bk)
        if not atoks or not btoks:
            continue
        inter = atoks & btoks
        union_ = atoks | btoks
        jac = len(inter) / len(union_) if union_ else 0
        dist_inter = distinctive(inter)
        smaller = atoks if len(atoks) <= len(btoks) else btoks
        larger = btoks if smaller is atoks else atoks
        is_subset = inter == smaller
        extra_a, extra_b = atoks - inter, btoks - inter
        extra_larger = larger - smaller

        if is_subset:
            # one name is a strict prefix/suffix of the other -- typically just
            # a verbosity difference (e.g. a tenant's name prepended to the
            # owner's field name). Safe UNLESS the extra includes a differing
            # lot NUMBER ("...Chemin 2" vs "...Chemin"), which does mean a
            # different specific plot despite the shared words.
            safe = not any(has_digit(t) for t in extra_larger)
        else:
            # neither name contains the other (e.g. "...Cañada Espartillo" vs
            # "...Cañada Norte", or lot "10" vs "19" of the same estación) --
            # only trust the overlap when whatever EACH side adds beyond it is
            # generic filler. Any digit, or any real distinctive word, left
            # over on either side means these are two different specific
            # places, not the same one described two ways.
            safe = (not any(has_digit(t) for t in extra_a) and not any(has_digit(t) for t in extra_b)
                    and not distinctive(extra_a) and not distinctive(extra_b))

        strong = jac >= 0.5 and is_subset and safe
        matched = False
        # non-subset case: even a single shared token is trustworthy as long
        # as it's genuinely distinctive (dist_inter, checked above) AND both
        # sides' remainders are pure filler (safe) -- e.g. "Nino Norte" vs
        # "Nino Sur" share only "Nino", but Norte/Sur carry zero risk of
        # meaning a different place, so one shared rare word is enough.
        if dist_inter and (strong or (safe and not is_subset)):
            matched = True
        elif (inter and ("COL" in atoks or "COL" in btoks) and same_operator(ak[1], bk[1])
              and inter == smaller
              and any(has_digit(t) for t in inter if t != "COL")
              and "LAS HERMANAS" not in ak[1] and "LAS HERMANAS" not in bk[1]):
            # "COL <n>" is Molisol/AFA/AgLab's own shorthand for Banchio's "Campo
            # Col" lots -- safe to match on the bare lot number here, but nowhere
            # else, since a bare number is otherwise ambiguous across this
            # operator's many differently-named parcels (Angel 3, Levrino, Las
            # Hermanas' own same-numbered lots, etc.). Requiring the matched
            # number itself (not just the word "COL") rules out a lone "COL"
            # entry from acting as a hub that merges every COL-N lot together.
            matched = True
        if matched:
            uf.union(ak, bk)
            merge_log.append((ak, bk))

print("within-dataset merges applied:", len(merge_log))
for a, b in merge_log:
    print(" ", a, "<->", b)

merged_groups = defaultdict(list)
for k, member_rows in groups.items():
    root = uf.find(k)
    merged_groups[root].extend(member_rows)

print("final campo groups (post-merge):", len(merged_groups))

# ---------- BUILD FIELD GROUPS ----------
def best_label(rowlist, key):
    c = Counter(r[key] for r in rowlist if r.get(key))
    if not c: return None
    return c.most_common(1)[0][0]

campos = []
gid = 0
for root, member_rows in merged_groups.items():
    gid += 1
    dataset = root[0]
    productor_disp = best_label(member_rows, "productor") or root[1].title()
    lote_disp = best_label(member_rows, "lote") or root[2].title()
    rows_sorted = sorted(member_rows, key=lambda r: (r.get("fecha") or "0000"))
    tipos_present = sorted(set(r["tipo"] for r in member_rows if r.get("tipo")))
    labs_present = sorted(set(r["lab"] for r in member_rows if r.get("lab")))
    fechas = [r["fecha"] for r in member_rows if r.get("fecha")]
    campos.append({
        "id": f"g{gid}",
        "dataset": dataset,
        "productor": productor_disp,
        "lote": lote_disp,
        "tipos": tipos_present,
        "labs": labs_present,
        "n_analyses": len(member_rows),
        "first_date": min(fechas) if fechas else None,
        "last_date": max(fechas) if fechas else None,
        "rows": rows_sorted,
    })

# ---------- PARAM METADATA ----------
CHEM_PARAM_META = {
    "MO":        {"label": "Materia Orgánica", "unit": "%", "dir": "hi", "crit": 2.0, "opt": 3.5},
    "C_Org":     {"label": "Carbono Orgánico", "unit": "%", "dir": "hi", "crit": 1.0, "opt": 1.8},
    "N_Total":   {"label": "Nitrógeno Total", "unit": "%", "dir": "hi", "crit": 0.10, "opt": 0.15},
    "pH":        {"label": "pH", "unit": "", "dir": "rng", "range_opt": [6.0, 7.0], "range_warn": [5.5, 7.5]},
    "CE":        {"label": "Conductividad Eléctrica", "unit": "µS/cm", "dir": "lo", "crit": 800, "opt": 400},
    "P_Total":   {"label": "Fósforo Total", "unit": "ppm", "dir": "hi", "crit": 20, "opt": 40},
    "P_BrayI":   {"label": "Fósforo Bray I", "unit": "ppm", "dir": "hi", "crit": 15, "opt": 30},
    "N_NO3":     {"label": "Nitrógeno de Nitratos", "unit": "ppm", "dir": "hi", "crit": 5, "opt": 15},
    "Azufre_S":  {"label": "Azufre (sulfatos)", "unit": "ppm", "dir": "hi", "crit": 5, "opt": 10},
    "CIC":       {"label": "Capacidad Intercambio Catiónico", "unit": "meq/100g", "dir": "hi", "crit": 15, "opt": 25},
    "Ca_Interc": {"label": "Calcio Intercambiable", "unit": "ppm", "dir": "hi", "crit": 500, "opt": 1000},
    "Mg_Interc": {"label": "Magnesio Intercambiable", "unit": "ppm", "dir": "hi", "crit": 50, "opt": 150},
    "K_Interc":  {"label": "Potasio Intercambiable", "unit": "ppm", "dir": "hi", "crit": 100, "opt": 200},
    "Na_Interc": {"label": "Sodio Intercambiable", "unit": "ppm", "dir": "lo", "crit": 200, "opt": 100},
    "PSI":       {"label": "% Sodio Intercambiable", "unit": "%", "dir": "lo", "crit": 15, "opt": 10},
    "Sat_Bases": {"label": "Saturación de Bases", "unit": "%", "dir": "hi", "crit": 60, "opt": 80},
    "Zn":        {"label": "Zinc disponible", "unit": "ppm", "dir": "hi", "crit": 0.5, "opt": 1.0},
    "Mn":        {"label": "Manganeso disponible", "unit": "ppm", "dir": "hi", "crit": 1, "opt": 5},
    "Fe":        {"label": "Hierro disponible", "unit": "ppm", "dir": "hi", "crit": 4, "opt": 20},
    "Cu":        {"label": "Cobre disponible", "unit": "ppm", "dir": "hi", "crit": 0.2, "opt": 0.5},
    "Bo":        {"label": "Boro disponible", "unit": "ppm", "dir": "hi", "crit": 0.5, "opt": 1.0},
    "Agua_Util": {"label": "Agua Útil", "unit": "%", "dir": "hi", "crit": 15, "opt": 25},
}
BIO_PARAM_META = {
    "N_bio":  {"label": "Nitrógeno biodisponible", "unit": "ppm", "dir": "hi", "crit": 20, "opt": 40},
    "P_bio":  {"label": "Fósforo biodisponible", "unit": "ppm", "dir": "hi", "crit": 25, "opt": 50},
    "K_bio":  {"label": "Potasio biodisponible", "unit": "ppm", "dir": "hi", "crit": 60, "opt": 120},
    "pH":     {"label": "pH", "unit": "", "dir": "rng", "range_opt": [6.0, 7.5], "range_warn": [5.5, 8.0]},
    "CE":     {"label": "Conductividad Eléctrica", "unit": "µS/cm", "dir": "lo", "crit": 800, "opt": 400},
    "CO2_Resp": {"label": "CO₂ Respiración", "unit": "ppm", "dir": "hi", "crit": 1500, "opt": 2500},
    "Pct_Organico": {"label": "% Orgánico (cromatografía)", "unit": "%", "dir": "hi", "crit": 40, "opt": 65},
    "Pct_Mineral":  {"label": "% Mineral (cromatografía)", "unit": "%", "dir": "lo", "crit": 80, "opt": 50},
    "Pct_CentroGaseoso": {"label": "% Centro Gaseoso", "unit": "%", "dir": "hi", "crit": 1, "opt": 3},
    "ICS": {"label": "Índice de Calidad de Suelo (ICS)", "unit": "1-3", "dir": "hi", "crit": 1.5, "opt": 2.5},
    "Act_Biologica": {"label": "Actividad Biológica", "unit": "1-3", "dir": "hi", "crit": 1.5, "opt": 2.5},
    "Compactacion": {"label": "Compactación", "unit": "1-3", "dir": "hi", "crit": 1.5, "opt": 2.5},
    "Respiracion": {"label": "Respiración", "unit": "1-3", "dir": "hi", "crit": 1.5, "opt": 2.5},
}

out = {
    "campos": campos,
    "chem_param_meta": CHEM_PARAM_META,
    "bio_param_meta": BIO_PARAM_META,
    "stats": {
        "n_campos": len(campos),
        "n_campos_quimico": sum(1 for c in campos if c["dataset"] == "quimico"),
        "n_campos_biologico": sum(1 for c in campos if c["dataset"] == "biologico"),
        "n_chem": n_chem,
        "n_bio": n_bio,
        "n_productores": len(set(c["productor"] for c in campos)),
    }
}
with open(os.path.join(OUT_DIR, "data.json"), "w", encoding="utf-8") as f:
    json.dump(out, f, ensure_ascii=False)

print("wrote data.json, campos:", len(campos))
